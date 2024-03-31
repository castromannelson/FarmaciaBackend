from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import status
from .serializer import *
from .models import *
from openpyxl import Workbook
from django.http import HttpResponse


# Create your views here.
#Listar
@permission_classes([IsAuthenticated])
class MedicinesListView(APIView):
    def get(self,request):
        medicines = Medicine.objects.all()
        
        serializer = MedicineSerializer(medicines, many = True)
        
        return Response(serializer.data)

#Obtener
@permission_classes([IsAuthenticated])
class MedicineView(APIView):
    def get(self,request,id):
        try:
            medicine = Medicine.objects.get(id=id)
        except Medicine.DoesNotExist:
            return Response({'Error': 'Medicine not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = MedicineSerializer(medicine)
        return Response(serializer.data)
        
#Crear
@permission_classes([IsAuthenticated])
class CreateMedicineView(APIView):

    def post(self,request):
        serializer = MedicineSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#Editar
@permission_classes([IsAuthenticated])
class EditMedicineView(APIView):
    def put(self, request, id,):
        try:
            medicine = Medicine.objects.get(id=id)
        except Medicine.DoesNotExist:
            return Response({'Error': 'Medicine not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = MedicineSerializer(medicine,data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


#Eliminar
@permission_classes([IsAuthenticated])
class DeleteMedicineView(APIView):

    def delete(self,request,id):
        try:
            medicine = Medicine.objects.get(id=id)
        except Medicine.DoesNotExist:
            return Response({'Error': 'Medicine not found'}, status=status.HTTP_404_NOT_FOUND)
        
        medicine.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

#Generar Reporte de Medicamentos
class ReporteMedicineExel(APIView):
    def get(self,request,*args,**kwargs):
        medicines = Medicine.objects.all()
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Reporte de Medicamentos'
        
        ws.merge_cells('A1:E1')
        ws['A3'] = 'Nombre'
        ws['B3'] = 'Grupo'
        ws['C3'] = 'Existencia'
        ws['D3'] = 'Precio'
        ws['E3'] = 'Estado'
        
        cont = 4
        
        for medicine in medicines:
            ws.cell(row = cont, column = 1).value = medicine.nameMedicine
            ws.cell(row = cont, column = 2).value = medicine.group
            ws.cell(row = cont, column = 3).value = medicine.exist
            ws.cell(row = cont, column = 4).value = medicine.price
            ws.cell(row = cont, column = 5).value = medicine.state
            cont+=1
    
        nombre_archivo = "ReporteMedicamentosExel.xlsx"
        response = HttpResponse(content_type="application/ms-exel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        
        return response