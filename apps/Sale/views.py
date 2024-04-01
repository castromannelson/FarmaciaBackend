import datetime
from rest_framework.response import Response
from rest_framework.decorators import permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework import status
from .serializer import *
from .models import *
from datetime import datetime, timedelta
from django.http import HttpResponse
from openpyxl import Workbook

# Create your views here.
# Listar


@permission_classes([IsAuthenticated])
class SalesListView(APIView):
    def get(self, request):
        sales = Sale.objects.all()

        serializer = SaleSerializer(sales, many=True)

        return Response(serializer.data)

# Listar Filtrado


@permission_classes([IsAuthenticated])
class SalesFilterListView(APIView):
    def get(self, request):
        sales = Sale.objects.filter(
            saleDate__gte=(datetime.now()-timedelta(days=7)))

        serializer = SaleSerializer(sales, many=True)

        return Response(serializer.data)

# Obtener


@permission_classes([IsAuthenticated])
class SaleView(APIView):
    def get(self, request, id):
        try:
            sale = Sale.objects.get(id=id)
        except Sale.DoesNotExist:
            return Response({'Error': 'Sale not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SaleSerializer(sale)
        return Response(serializer.data)

# Crear


@permission_classes([IsAuthenticated])
class CreateSaleView(APIView):
    def post(self, request):
        book = DateBook.objects.get(batch=request.data.get('batch'))
        amount = request.data.get('amount')
        if ((book.amountBatch-int(amount)) > 0):
            serializer = SaleSerializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'Error': "No se puede realizar la venta"}, status=status.HTTP_400_BAD_REQUEST)

# Editar


@permission_classes([IsAuthenticated])
class EditSaleView(APIView):
    def put(self, request, id):
        try:
            sale = Sale.objects.get(id=id)
        except Sale.DoesNotExist:
            return Response({'Error': 'Sale not found'}, status=status.HTTP_404_NOT_FOUND)

        serializer = SaleSerializer(sale, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# Eliminar
@permission_classes([IsAuthenticated])
class DeleteSaleView(APIView):
    def delete(self, request, id):
        try:
            sale = Sale.objects.get(id=id)
        except Sale.DoesNotExist:
            return Response({'Error': 'Sale not found'}, status=status.HTTP_404_NOT_FOUND)

        sale.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# Generar Reporte de Ventas


class ReporteSaleExel(APIView):
    def get(self, request, *args, **kwargs):
        # Filtrar las ventas desde hace una semana hasta hoy
        sales = Sale.objects.filter(
            saleDate__gte=(datetime.now()-timedelta(days=7)))

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Factura de Ventas'

        ws.merge_cells('A1:E1')
        ws['A3'] = 'Medicamento'
        ws['B3'] = 'Cantidad'
        ws['C3'] = 'Fecha de Venta'
        ws['D3'] = 'Lote'
        ws['E3'] = 'Importe'
        cont = 4
        total = 0

        for sale in sales:
            ws.cell(row=cont, column=1).value = sale.medicine.nameMedicine
            ws.cell(row=cont, column=2).value = sale.amount
            ws.cell(row=cont, column=3).value = sale.saleDate
            ws.cell(row=cont, column=4).value = sale.batch.batch
            ws.cell(row=cont, column=5).value = sale.profit
            total += sale.profit
            cont += 1

        ws.cell(row=cont, column=5).value = "Total: " + str(total)

        # Obtener las fechas
        hoy = datetime.now()
        hace_una_semana = hoy - timedelta(days=7)

        # Formatear las fechas como strings
        fecha_actual_str = hoy.strftime("%d-%m-%Y")
        fecha_hace_una_semana_str = hace_una_semana.strftime("%d-%m-%Y")

        # Concatenar el nombre del archivo con las fechas
        nombre_archivo = f"Factura de Ventas de {
            fecha_hace_una_semana_str} a {fecha_actual_str}.xlsx"
        response = HttpResponse(content_type="application/ms-exel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)

        return response
